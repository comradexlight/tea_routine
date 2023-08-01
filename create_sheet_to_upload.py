from typing import List

from icu import Collator, Locale
from openpyxl import Workbook

from get_data_from_sheet import SaleRecordItem


def create_sheet_to_upload(wb2upload: Workbook,
                           list2upload: List[SaleRecordItem],
                           sheet_name: str) -> None:
    """
    The function inserts the data from the list to upload into the sheet of
    xlsx document to be uploaded.
    """
    new_ws = wb2upload.create_sheet(sheet_name)
    adjusted_width = 0
    _sort_list2upload(list2upload)
    for row, element in enumerate(list2upload, start=1):
        title = element.title.removesuffix(", , шт").removesuffix(", , кг")

        if adjusted_width < len(title):
            adjusted_width = len(title)

        new_ws.cell(row=row, column=1, value=title)
        new_ws.cell(row=row, column=2, value=element.price)
        new_ws.cell(row=row, column=3, value=element.qty)
        new_ws.cell(row=row, column=4, value=element.amount).number_format = "0"
        new_ws.cell(row=row, column=5, value=f"=1-D{row}/(B{row}*C{row})").number_format = '0.00%'

    new_ws.column_dimensions['A'].width = adjusted_width
    new_ws.cell(row=len(list2upload) + 2, column=4, value=f"=SUM(D1:D{len(list2upload)})")


def _sort_list2upload(list2upload: List[SaleRecordItem]) -> None:
    """
    The function sorts the items in the list of SaleRecordItems by title in ABC order.
    """
    collator = Collator.createInstance(Locale('ru_RU'))
    list2upload.sort(key=lambda element: collator.getSortKey(element.title))
