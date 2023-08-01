"""
A simple script to prepare xlsx files for upload to 1c.
"""
from time import time
from sys import argv

from openpyxl import Workbook
from openpyxl import load_workbook

from get_active_ws import get_active_ws
from get_data_from_sheet import get_data_from_sheet
from create_sheet_to_upload import create_sheet_to_upload


def create_xlsx_to_upload() -> None:
    """
    The main function of this script.
    """
    path = argv[1]
    input_wb = load_workbook(filename=path, read_only=False)
    output_wb = Workbook()
    active_ws = get_active_ws(input_wb)
    lists_to_upload = [get_data_from_sheet(input_ws) for input_ws in active_ws]
    [create_sheet_to_upload(output_wb, data_list, new_ws.title) for new_ws,
            data_list in zip(active_ws, lists_to_upload)]
    output_wb.save("2upload.xlsx")


def main() -> None:
    """
    The Entry point.
    """
    create_xlsx_to_upload()


if __name__ == '__main__':
    start_time = time()
    main()
    print(time() - start_time)
